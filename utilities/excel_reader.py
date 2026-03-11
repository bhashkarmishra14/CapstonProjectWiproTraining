from openpyxl import load_workbook


# Reads test rows from Excel sheet and returns list of dictionaries.
def read_test_data(file_path, sheet_name="Sheet1"):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    headers = [cell.value for cell in sheet[1]]
    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        rows.append(dict(zip(headers, row)))

    workbook.close()
    return rows
